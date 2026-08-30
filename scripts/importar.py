#!/usr/bin/env python3
# Importa la flota real a Supabase.
# Fuente principal: ControlCobros (estado financiero actual, limpio).
# Enriquecimiento: Hoja de vida (km actual por carro).
#
# Uso:
#   python3 scripts/importar.py            # DRY-RUN (no escribe, solo muestra el plan)
#   python3 scripts/importar.py --commit   # escribe a Supabase (idempotente)
import os, re, sys, json, warnings
warnings.filterwarnings("ignore")
import openpyxl, requests

COMMIT = "--commit" in sys.argv
CONTROL = "/Users/mateozuluagatoledo/Downloads/ControlCobros_20260827 tarde.xlsx"
HOJAVIDA = "/Users/mateozuluagatoledo/Desktop/CLAUDE PROJECTS/autolujo/Hoja de vida carros Autolujo Actualizadas 2025.xlsx"

# --- credenciales desde .env.local ---
env = {}
for line in open(os.path.join(os.path.dirname(__file__), "..", ".env.local")):
    line = line.strip()
    if "=" in line and not line.startswith("#"):
        k, v = line.split("=", 1); env[k] = v
URL = env["NEXT_PUBLIC_SUPABASE_URL"]; KEY = env["SUPABASE_SERVICE_ROLE_KEY"]
H = {"apikey": KEY, "Authorization": f"Bearer {KEY}", "Content-Type": "application/json",
     "Prefer": "return=representation"}

def sb_get(path):
    r = requests.get(f"{URL}/rest/v1/{path}", headers=H); r.raise_for_status(); return r.json()
def sb_post(table, rows):
    if not rows: return []
    r = requests.post(f"{URL}/rest/v1/{table}", headers=H, json=rows)
    if r.status_code >= 300: print("  ERR", table, r.status_code, r.text[:300]); return []
    return r.json()

def canon(s):
    t = re.sub(r"[^A-Z0-9]", "", str(s).upper())
    m = re.match(r"^([A-Z]*)0*(\d+)$", t)
    return (m.group(1) + str(int(m.group(2)))) if m else t

def empresa_codigo(s):
    s = str(s).lower()
    if "exclient" in s: return None
    if "kowua" in s: return "KOWUA"
    if "gold" in s: return "GOLD"
    if "lujo" in s or "auto" in s: return "AUTOLUJO"
    return None

def num(v):
    try: return float(v)
    except: return 0.0

# ---------- 1) Parsear ControlCobros ----------
wb = openpyxl.load_workbook(CONTROL, data_only=True, read_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
idx = {h.split(",")[0]: i for i, h in enumerate(rows[0]) if h}
def C(r, name):
    i = idx.get(name); return r[i] if i is not None and i < len(r) else None

clientes = {}   # codigo -> {codigo,nombre,telefono,whatsapp}
vehiculos = {}  # (emp,numero) -> {emp,numero}
contratos = []  # {emp,numero,cliente_codigo,letra,abono,saldo,fecha,sinies,otras}
saltados = 0
for r in rows[1:]:
    if str(C(r, "FORMATO")) != "1": continue  # solo filas de contrato
    emp = empresa_codigo(C(r, "EMPRESA"))
    unidad = C(r, "UNIDAD"); code = C(r, "CONDUCTOR")
    if not emp or not unidad:
        saltados += 1; continue
    numero = str(unidad).strip()
    code = str(code).strip() if code else None
    if code:
        clientes.setdefault(code, {
            "codigo": code,
            "nombre": (str(C(r, "NOMBRE")).strip() if C(r, "NOMBRE") else "Sin nombre"),
            "telefono": (str(C(r, "TELEFONO")).strip() if C(r, "TELEFONO") else None),
        })
    vehiculos.setdefault((emp, numero), {"emp": emp, "numero": numero})
    fec = C(r, "FEC_INGRES")
    contratos.append({
        "emp": emp, "numero": numero, "cliente_codigo": code,
        "letra": num(C(r, "CUOTA")), "abono": num(C(r, "FON_INSCRI")),
        "saldo": num(C(r, "DEU_RENTA")),
        "fecha": (str(fec)[:10] if fec else None),
        "sinies": num(C(r, "DEU_SINIES")), "otras": num(C(r, "DEU_OTRAS")),
    })

# ---------- 2) Km desde Hoja de vida (best-effort) ----------
km_por_carro = {}   # canon(numero) -> km
try:
    wv = openpyxl.load_workbook(HOJAVIDA, data_only=True, read_only=True)
    for name in wv.sheetnames:
        m = re.search(r"CARRO\s+([A-Z]?-?\d+)", name.upper())
        if not m: continue
        key = canon(m.group(1))
        ws2 = wv[name]; ultimo = None
        for r in ws2.iter_rows(min_row=3, max_col=3, values_only=True):
            v = r[2] if len(r) > 2 else None
            if v is not None and str(v).replace(".", "").isdigit():
                ultimo = int(float(v))
        if ultimo: km_por_carro[key] = ultimo
except Exception as e:
    print("Aviso: no pude leer km de Hoja de vida:", e)

# ---------- Resumen ----------
print(f"\n{'='*60}\n{'DRY-RUN (no escribe)' if not COMMIT else 'ESCRIBIENDO A SUPABASE'}\n{'='*60}")
print(f"ControlCobros: {len(contratos)} contratos activos | filas saltadas (exclientes/vacías): {saltados}")
from collections import Counter
print("Por empresa:", dict(Counter(c['emp'] for c in contratos)))
print(f"Clientes únicos: {len(clientes)} | Vehículos únicos: {len(vehiculos)}")
print(f"Km encontrados en Hoja de vida: {len(km_por_carro)} carros")
print("\nMuestra de 3 contratos:")
for c in contratos[:3]:
    kk = km_por_carro.get(canon(c["numero"]))
    print(f"  {c['emp']} #{c['numero']} · letra ${c['letra']:.0f} · saldo ${c['saldo']:.0f} · abono ${c['abono']:.0f} · km {kk or '—'} · desde {c['fecha']}")

if not COMMIT:
    print("\n👉 Revisa los números. Si están bien, corre con --commit para escribir.")
    sys.exit(0)

# ---------- 3) Escribir (idempotente) ----------
emp_id = {e["codigo"]: e["id"] for e in sb_get("empresas?select=id,codigo")}

# clientes
exist_cli = {c["codigo"]: c["id"] for c in sb_get("clientes?select=id,codigo") if c.get("codigo")}
nuevos_cli = [c for k, c in clientes.items() if k not in exist_cli]
for c in nuevos_cli: c["whatsapp"] = ("+507" + re.sub(r"\D", "", c["telefono"])) if c.get("telefono") else None
creados = sb_post("clientes", [{"codigo": c["codigo"], "nombre": c["nombre"], "telefono": c.get("telefono"), "whatsapp": c.get("whatsapp")} for c in nuevos_cli])
for c in creados: exist_cli[c["codigo"]] = c["id"]
print(f"Clientes: +{len(creados)} nuevos ({len(exist_cli)} total)")

# vehiculos
exist_veh = {(v["empresa_id"], v["numero"]): v["id"] for v in sb_get("vehiculos?select=id,numero,empresa_id")}
nuevos_veh = []
for (emp, numero), v in vehiculos.items():
    eid = emp_id.get(emp)
    if eid and (eid, numero) not in exist_veh:
        nuevos_veh.append({"empresa_id": eid, "numero": numero, "km_actual": km_por_carro.get(canon(numero), 0)})
creados = sb_post("vehiculos", nuevos_veh)
for v in creados: exist_veh[(v["empresa_id"], v["numero"])] = v["id"]
print(f"Vehículos: +{len(creados)} nuevos ({len(exist_veh)} total)")

# contratos (uno activo por vehículo)
con_activo = {c["vehiculo_id"] for c in sb_get("contratos?select=vehiculo_id&estado=eq.activo")}
nuevos_con = []
cargos = []
for c in contratos:
    eid = emp_id.get(c["emp"]); vid = exist_veh.get((eid, c["numero"])); cid = exist_cli.get(c["cliente_codigo"])
    if not (eid and vid and cid): continue
    if vid in con_activo: continue
    con_activo.add(vid)
    fila = {"vehiculo_id": vid, "cliente_id": cid, "empresa_id": eid,
            "letra_diaria": c["letra"] or 0, "abono_inicial": c["abono"] or 0,
            "saldo_inicial": c["saldo"] or 0, "estado": "activo"}
    if c["fecha"]: fila["fecha_inicio"] = c["fecha"]
    nuevos_con.append((fila, c))
creados = sb_post("contratos", [f for f, _ in nuevos_con])
print(f"Contratos: +{len(creados)} nuevos")
# cargos por siniestros/otras del contrato recién creado
for cont, (_, c) in zip(creados, nuevos_con):
    if c["sinies"] > 0: cargos.append({"contrato_id": cont["id"], "tipo": "siniestro", "concepto": "Saldo siniestros (migrado)", "monto": c["sinies"]})
    if c["otras"] > 0: cargos.append({"contrato_id": cont["id"], "tipo": "otras", "concepto": "Otras deudas (migrado)", "monto": c["otras"]})
sb_post("cargos", cargos)
print(f"Cargos (siniestros/otras): +{len(cargos)}")
print("\n✅ Importación completa.")
